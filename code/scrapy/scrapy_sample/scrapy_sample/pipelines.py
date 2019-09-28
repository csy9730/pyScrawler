# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
""" 
    json,jsonline,markdown,xls
    数据库：PostgreSQL,sqlite,pymongo
"""

from scrapy import signals, Request
from scrapy.pipelines.images import ImagesPipeline
from scrapy.http import Request
from scrapy_sample.items import ImageItem
from scrapy.exporters import JsonItemExporter, JsonLinesItemExporter

import os
import json
import codecs
import sqlite3
import openpyxl
import logging
_log = logging.getLogger(__name__)

class RawFilenameImagePipeline(ImagesPipeline):
    """
        add referer,set img_folder,img_name
    """
    def get_media_requests(self, item, info):
        if not isinstance(item, ImageItem):
            return
        requests = super().get_media_requests(item, info)
        
        for req in requests:
            req.headers["referer"]=item['referer']
            req.meta["img_folder"] = item["img_folder"] or ''             
        return requests

        # for image_url in item['image_urls']:
        #    yield Request(url=image_url,headers={'Referer':item['header_referer']})
    def file_path(self, request, response=None, info=None):        
        url = request.url   
        return 'full/{1}{0}'.format( url.split('/')[-1],  request.meta["img_folder"])

        # image_guid = hashlib.sha1(to_bytes(request.url)).hexdigest()
        # return 'full/%s.jpg' % (image_guid)
    
class RefererImagePipeline(ImagesPipeline):
    def get_media_requests(self, item, info):
        requests = super().get_media_requests(item, info)
        for req in requests:
            req.headers.appendlist("referer", item['referer'])
        return requests

class CsdnBlogBackupPipeline(object):
    def process_item(self, item, spider):
        dirname = 'blogs'

        if not os.path.exists(dirname):
            os.mkdir(dirname)
        print(r''+dirname+os.sep+item["title"]+".md")
        with codecs.open(f'{dirname}{os.sep}{item["title"]}.md', 'w', encoding='utf-8') as f:
            f.write(item['content'])
            f.close()
        return item

class JsonExportPipeline(object):
    def __init__(self):
        _log.info('JsonExportPipeline.init....')
        self.files = {}

    @classmethod
    def from_crawler(cls, crawler):
        _log.info('JsonExportPipeline.from_crawler....')
        pipeline = cls()
        crawler.signals.connect(pipeline.spider_opened, signals.spider_opened)
        crawler.signals.connect(pipeline.spider_closed, signals.spider_closed)
        return pipeline

    def spider_opened(self, spider):
        _log.info('JsonExportPipeline.spider_opened....')
        file = open('%s.json' % spider.name, 'w+b')
        self.files[spider] = file
        self.exporter = JsonItemExporter(file)
        self.exporter.start_exporting()

    def spider_closed(self, spider):
        _log.info('JsonExportPipeline.spider_closed....')
        self.exporter.finish_exporting()
        file = self.files.pop(spider)
        file.close()

    def process_item(self, item, spider):
        _log.info('JsonExportPipeline.process_item....')
        self.exporter.export_item(item)
        return item
class JsonlineExportPipeline(object):
    def __init__(self):
        _log.info('JsonExportPipeline.init....')
        self.files = {}

    @classmethod
    def from_crawler(cls, crawler):
        _log.info('JsonlineExportPipeline.from_crawler....')
        pipeline = cls()
        crawler.signals.connect(pipeline.spider_opened, signals.spider_opened)
        crawler.signals.connect(pipeline.spider_closed, signals.spider_closed)
        return pipeline

    def spider_opened(self, spider):
        _log.info('JsonExportPipeline.spider_opened....')
        file = open('%s.json' % spider.name, 'w+b')
        self.files[spider] = file
        self.exporter = JsonLinesItemExporter(file)
        self.exporter.start_exporting()

    def spider_closed(self, spider):
        _log.info('JsonExportPipeline.spider_closed....')
        self.exporter.finish_exporting()
        file = self.files.pop(spider)
        file.close()

    def process_item(self, item, spider):
        _log.info('JsonExportPipeline.process_item....')
        self.exporter.export_item(item)
        return item
    # sqlite

class xlsPipeline(object):
    #打开数据库
    def open_spider(self, spider):        
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.get_active_sheet()
        self.xls_name = "scr_scrapy.xlsx"
        if os.path.exists( self.xls_name ):
            self.wb = openpyxl.load_workbook(self.xls_name)            
            sheetnames = self.wb.get_sheet_names()# 取第一张表
            self.sheet = self.wb.get_sheet_by_name(sheetnames[0])
    #关闭数据库
    def close_spider(self, spider):
        self.wb.save(self.xls_name)
    #对数据进行处理
    def process_item(self, item, spider): 
        self.sheet.append( [json.dumps( item[key],ensure_ascii=False )  for key  in item.keys() ] )
        # self.sheet.append( [item["image_urls"][0],item["title"], ] )
        return item


class SQLiteImagePipeline(object):
    #打开数据库
    def open_spider(self, spider):
        db_name = spider.settings.get('SQLITE_DB_NAME', 'scr_image.db')       
        if not (os.path.exists(db_name) and os.path.isfile(db_name)):
            self.db_conn = sqlite3.connect(db_name)
            self._new_table()
        else:
            self.db_conn = sqlite3.connect(db_name)      
        self.db_cur = self.db_conn.cursor()

    #关闭数据库
    def close_spider(self, spider):
        self.db_conn.commit()
        self.db_conn.close()

    #对数据进行处理
    def process_item(self, item, spider):
        if not isinstance(item, ImageItem):
            return item
        self._insert_db(item)
        return item
    def _new_table(self):
        self.db_conn.execute('''CREATE TABLE IMAGES
            (ID CHAR(32) PRIMARY KEY     NOT NULL,            
            image_url       CHAR(150),
            REFERER        CHAR(150),
            path       CHAR(150),
            title       CHAR(150));''')  
    #插入数据
    def _insert_db(self, item):
        imgs = item['images']
        for img in imgs:
            if img["checksum"] is not None:
                values = (
                    img["checksum"],
                    img["url"],
                    item['referer'],
                    img["path"].encode('utf8'),                    
                    item['title'].encode('utf8'),
                )
                sql = 'INSERT INTO IMAGES VALUES(?,?,?,?,?)'
                self.db_cur.execute(sql, values)
# import pymongo
class BudejieMongoPipeline(object):
    "将百思不得姐段子保存到MongoDB中"
    collection_name = 'jokes'

    def __init__(self, mongo_uri, mongo_db):
        self.mongo_uri = mongo_uri
        self.mongo_db = mongo_db

    @classmethod
    def from_crawler(cls, crawler):
        return cls(
            mongo_uri=crawler.settings.get('MONGO_URI'),
            mongo_db=crawler.settings.get('MONGO_DATABASE', 'budejie')
        )

    def open_spider(self, spider):
        self.client = pymongo.MongoClient(self.mongo_uri)
        self.db = self.client[self.mongo_db]

    def close_spider(self, spider):
        self.client.close()

    def process_item(self, item, spider):
        self.db[self.collection_name].insert_one(dict(item))
        return item

# import psycopg2
class BudejiePostgrePipeline(object):
    "将百思不得姐段子保存到PostgreSQL中"

    def __init__(self):
        self.connection = psycopg2.connect("dbname='test' user='postgres' password='12345678'")
        self.connection.autocommit = True

    def open_spider(self, spider):
        self.cursor = self.connection.cursor()

    def close_spider(self, spider):
        self.cursor.close()
        self.connection.close()

    def process_item(self, item, spider):
        self.cursor.execute('insert into joke(author,content) values(%s,%s)', (item['username'], item['content']))
        return item

