import json
import re,os,sys
import requests
def ssdf():
    pass
    #    print(vars(settings))
#    settings.set("CLOSESPIDER_ITEMCOUNT", 3, priority='cmdline')
def a():
    comments = requests.get('http://comment5.news.sina.com.cn/page/info?version=1&format=js&channel=gn&newsid=comos-fyfzhac1650783')
    comments.encoding = 'utf-8'
    print(comments)
    jd = json.loads(comments.text.strip('var data=')) #移除改var data=将其变为json数据
    print(jd['result']['count']['total'])
def jsonRegexDemo():
    st = """var comicName = "琅琊榜";var nextVolume="#";var preVolume="#";var picCount = 9;var picAy = new Array();var hosts = ["http://coldpic.sfacg.com","http://coldpic.sfacg.com", "http://ltpic.sfacg.com"];picAy[0] = "/Pic/OnlineComic4/LYB/ZP/0003_7815/001_764.jpg";picAy[1] = "/Pic/OnlineComic4/LYB/ZP/0003_7815/002_761.jpg";picAy[2] = "/Pic/OnlineComic4/LYB/ZP/0003_7815/003_340.jpg";picAy[3] = "/Pic/OnlineComic4/LYB/ZP/0003_7815/004_581.jpg";picAy[4] = "/Pic/OnlineComic4/LYB/ZP/0003_7815/005_985.jpg";picAy[5] = "/Pic/OnlineComic4/LYB/ZP/0003_7815/006_672.jpg";picAy[6] = "/Pic/OnlineComic4/LYB/ZP/0003_7815/007_272.jpg";picAy[7] = "/Pic/OnlineComic4/LYB/ZP/0003_7815/008_851.jpg";picAy[8] = "/Pic/OnlineComic4/LYB/ZP/0003_7815/009_256.jpg";"""
    print(st)
    st2 = st.split(';')
    print(st2)
    imgUrl = re.findall(r"=\s\"([\w\d_/]+.jpg\")",st)
    print(imgUrl,len(imgUrl))
def excelDemo():
    import openpyxl      
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    xls_name = "scr_scrapy.xlsx"
    if os.path.exists( xls_name ):
        wb = openpyxl.load_workbook(xls_name)
        sheetnames = wb.get_sheet_names()
        sheet = wb.get_sheet_by_name(sheetnames[0])
        item = {"abc":456,"def":"wrtretert"}
        sheet.append( [json.dumps( item[key],ensure_ascii=False )  for key  in item.keys() ] )
        wb.save(xls_name)

def main():
    pass
    excelDemo()
    #jsonRegexDemo()
def execJsDemo():
    import execjs
    import os
    os.environ["EXECJS_RUNTIME"] = 'Node'
    print( execjs.get().name )
    # 执行 JS 语句
    ctx = execjs.compile(""" 
    function add(x, y) {
       return x + y;
    }
    """)
    print( ctx.call("add", 1, 2) )
    return 
    with open('./test.js') as f:  # 执行 JS 文件
        ctx = execjs.compile(f.read())
        ctx.call('add', 1, 2)
def addLink(self,response):
    xp = '//div[@class="pagination"]//a[@class="next_page"]/@href'
    rules = [ {"selector":"xpath","path":"","re":".*","callback":None,"format":"https://konachan.com{0}"}]
    for r in rules:
        urls = response.xpath(r["path"]).getall()
        for url in urls:
            yield  Request( r["format"].format(url) ,callback=self.parse_image)
def parse(self,response):
    # tart url 中 分配  callback_index  # url = response.url
    callback_index = response.meta["response"]
    resp = bg[callback_index]
    resp["rule"]
    resp[""]


if __name__ == "__main__":
    main()

